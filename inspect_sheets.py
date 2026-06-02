"""
ตรวจสอบโครงสร้าง Excel — แสดงชื่อ sheet และ headers ของแต่ละ sheet
รัน:  python inspect_sheets.py   (หรือ py inspect_sheets.py)
"""
import os, json
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "KPI Ticket Resolution.xlsx")
OUT  = os.path.join(HERE, "sheet_inspection.json")

wb = load_workbook(XLSX, data_only=True, read_only=True)
report = {}
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}\n")

for name in wb.sheetnames:
    ws = wb[name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = []
    sample_rows = []
    try:
        first = next(rows_iter)
        headers = [str(h) if h is not None else "(empty)" for h in first]
        # collect first 3 data rows for context
        for i, r in enumerate(rows_iter):
            if i >= 3: break
            sample_rows.append([str(v) if v is not None else "" for v in r])
    except StopIteration:
        pass

    report[name] = {"headers": headers, "sample_rows": sample_rows}
    print(f"=== Sheet: '{name}' ===")
    print(f"  Headers ({len(headers)}): {headers}")
    if sample_rows:
        print(f"  Sample row 1: {sample_rows[0][:8]}{'...' if len(sample_rows[0]) > 8 else ''}")
    print()

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(report, f, ensure_ascii=False, indent=2)

print(f"✓ Written → {OUT}")
