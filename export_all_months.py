"""
Export ทุก sheet ที่เป็นข้อมูลดิบใน KPI Ticket Resolution.xlsx เป็น JSON ก้อนเดียว
รองรับ drill-down 4 มิติ: Team / Owner / Category / SubCategory
รัน:  python export_all_months.py
ผลลัพธ์:  all_months.json
"""
import json, os, statistics
from openpyxl import load_workbook
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "KPI Ticket Resolution.xlsx")
OUT  = os.path.join(HERE, "all_months.json")

REQUIRED_BASE = {"TeamOwner", "TicketOwner", "ResolutionTime", "FirstResponseTime"}

def num(v):
    if v is None or v == "" or v == "NULL": return None
    try: return float(v)
    except: return None

def s(v, default="(ไม่ระบุ)"):
    if v is None or v == "" or v == "NULL": return default
    return str(v).strip()

def agg(rows, dims):
    """dims = list of keys, returns aggregated rows by tuple of dims"""
    buckets = defaultdict(list)
    for r in rows:
        key = tuple(r[d] for d in dims)
        buckets[key].append(r)
    out = []
    for key, items in buckets.items():
        rv = [x["res"] for x in items if x["res"] is not None]
        fv = [x["frt"] for x in items if x["frt"] is not None]
        sv = [x["sla"] for x in items if x["sla"] is not None]
        row = {}
        for i, d in enumerate(dims):
            row[d_to_col(d)] = key[i]
        row["tickets"] = len(items)
        row["avgResolution"]    = round(statistics.mean(rv), 2) if rv else 0
        row["avgFirstResponse"] = round(statistics.mean(fv), 2) if fv else 0
        row["avgSLAGap"]        = round(statistics.mean(sv), 2) if sv else 0
        row["minResolution"]    = round(min(rv), 2) if rv else 0
        row["maxResolution"]    = round(max(rv), 2) if rv else 0
        out.append(row)
    return out

def d_to_col(d):
    return {
        "team": "TeamOwner",
        "owner": "TicketOwner",
        "category": "Category",
        "subcategory": "SubCategory",
    }[d]

wb = load_workbook(XLSX, data_only=True, read_only=True)
print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}\n")

months = {}
sheet_order = []
ai_workforce = []   # เก็บ AI Workforce data แยก

for name in wb.sheetnames:
    ws = wb[name]

    # ─── AI Workforce sheet (โครงสร้างต่างจาก month sheets) ───────────────
    if name == 'AI Workforce':
        rows_iter = list(ws.iter_rows(values_only=True))
        # Row 0 = title, Row 1 = subtitle, Row 2 = empty, Row 3 = headers
        header_row = None
        data_start = 0
        for ri, row in enumerate(rows_iter):
            vals = [str(v).strip() if v else '' for v in row]
            if 'ทีม' in vals:
                header_row = vals
                data_start = ri + 1
                break
        if header_row is None:
            print(f"  [Skip] '{name}': no header row found")
            continue
        # map header -> index
        hmap = {h: i for i, h in enumerate(header_row) if h}
        for row in rows_iter[data_start:]:
            if row is None or all(v is None for v in row): continue
            def cell(col):
                idx = hmap.get(col)
                v = row[idx] if idx is not None and idx < len(row) else None
                return str(v).strip() if v else ''
            team = cell('ทีม')
            if not team: continue
            ai_workforce.append({
                'team':      team,
                'stats':     cell('\U0001f4ca สถิติ 3 เดือน (Mar\u2013May)'),
                'painpoints':cell('\U0001f534 Pain Points หลัก'),
                'usecase':   cell('\U0001f916 AI Use Case ที่แนะนำ'),
                'howworks':  cell('\u2699\ufe0f วิธีทำงาน (How it works)'),
                'benefits':  cell('\u2705 ผลที่คาดว่าจะได้รับ'),
                'tools':     cell('\U0001f6e0\ufe0f เครื่องมือแนะนำ'),
                'priority':  cell('Priority'),
            })
        print(f"  [OK] '{name}': {len(ai_workforce)} teams")
        continue

    # ─── Monthly Ticket sheets ─────────────────────────────────────────────
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        continue

    header_map = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}
    has_sla = "SLAGap" in header_map or "DateRequiredRang" in header_map
    if not REQUIRED_BASE.issubset(header_map.keys()) or not has_sla:
        print(f"  [Skip] '{name}': not a data sheet")
        continue

    i_team  = header_map["TeamOwner"]
    i_owner = header_map["TicketOwner"]
    i_res   = header_map["ResolutionTime"]
    i_frt   = header_map["FirstResponseTime"]
    i_sla   = header_map["SLAGap"] if "SLAGap" in header_map else header_map["DateRequiredRang"]
    i_cat   = header_map.get("Category")
    i_sub   = header_map.get("SubCategory") or header_map.get("SubCatagory")

    rows = []
    for r in rows_iter:
        if r is None or all(v is None for v in r): continue
        team = r[i_team]
        if team is None: continue
        rec = {
            "team":  s(team),
            "owner": s(r[i_owner]) if i_owner is not None else "(ไม่ระบุ)",
            "category":    s(r[i_cat]) if i_cat is not None else "(ไม่ระบุ)",
            "subcategory": s(r[i_sub]) if i_sub is not None else "(ไม่ระบุ)",
            "res": num(r[i_res]),
            "frt": num(r[i_frt]),
            "sla": num(r[i_sla]),
        }
        rows.append(rec)

    if not rows: continue

    rv = [x["res"] for x in rows if x["res"] is not None]
    fv = [x["frt"] for x in rows if x["frt"] is not None]
    sv = [x["sla"] for x in rows if x["sla"] is not None]

    months[name] = {
        "overall": {
            "totalTickets": len(rows),
            "avgResolution":    round(statistics.mean(rv), 2) if rv else 0,
            "avgFirstResponse": round(statistics.mean(fv), 2) if fv else 0,
            "avgSLAGap":        round(statistics.mean(sv), 2) if sv else 0,
        },
        "team":        agg(rows, ["team"]),
        "owner":       agg(rows, ["team", "owner"]),
        "category":    agg(rows, ["category"]),
        "subcategory": agg(rows, ["category", "subcategory"]),
        "teamCategory":      agg(rows, ["team", "category"]),
        "teamSubcategory":   agg(rows, ["team", "category", "subcategory"]),
        "ownerCategory":     agg(rows, ["team", "owner", "category"]),
        "categorySubcategory": agg(rows, ["category", "subcategory"]),
        "raw": rows
    }
    sheet_order.append(name)
    print(f"  [OK] '{name}': {len(rows)} tickets, "
          f"{len(months[name]['team'])} teams, "
          f"{len(months[name]['owner'])} owners, "
          f"{len(months[name]['category'])} categories, "
          f"{len(months[name]['subcategory'])} subcategories")

with open(OUT, "w", encoding="utf-8") as f:
    json.dump({"months": months, "sheetOrder": sheet_order, "aiWorkforce": ai_workforce}, f, ensure_ascii=False, indent=2)

print(f"\n[Written] -> {OUT}")
print(f"  Months: {sheet_order}")
print(f"  AI Workforce rows: {len(ai_workforce)}")
